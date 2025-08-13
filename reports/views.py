from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from .forms import TaskSubmissionFormSet
from .models import TaskSubmission, Engineer, InventoryItem
from django.db.models import Count, Q
from django.http import HttpResponse
import pandas as pd
from weasyprint import HTML
import io
from django.utils import timezone
from datetime import datetime
from django.contrib.auth.decorators import user_passes_test
import os

def team_leader_required(view_func):
    def _wrapped_view(request, *args, **kwargs):
        if not hasattr(request.user, 'engineer') or not request.user.engineer.is_team_leader:
            return HttpResponse("You do not have permission to access this page.", status=403)
        return view_func(request, *args, **kwargs)
    return _wrapped_view

@login_required
def submit_tasks(request):
    if request.method == 'POST':
        formset = TaskSubmissionFormSet(request.POST)
        if formset.is_valid():
            saved_tasks = []
            for form in formset.forms:
                if form.cleaned_data and not form.cleaned_data.get('DELETE', False):
                    task = form.save(commit=False)
                    if hasattr(request.user, 'engineer') and request.user.engineer:
                        task.engineer = request.user.engineer
                        if not task.reporter:
                            task.reporter = request.user.engineer.name
                        if not task.date:
                            task.date = timezone.now().date()
                        task.save()
                        form.save_m2m()
                        # If reporter is blank or equals the primary engineer, append team members to it
                        if hasattr(request.user, 'engineer'):
                            names = [request.user.engineer.name] + [m.name for m in task.team_members.all()]
                            if not task.reporter or task.reporter.strip() == request.user.engineer.name.strip():
                                task.reporter = ", ".join(names)
                                task.save(update_fields=['reporter'])
                        saved_tasks.append(task)
            if saved_tasks:
                return redirect('reports:submission_confirmation')
    else:
        initial = []
        if hasattr(request.user, 'engineer'):
            initial.append({
                'date': timezone.now().date(),
                'reporter': request.user.engineer.name,
            })
        formset = TaskSubmissionFormSet(initial=initial or None)
    return render(request, 'submit_tasks.html', {'formset': formset})

@login_required
def submission_confirmation(request):
    return render(request, 'submission_confirmation.html')

@team_leader_required
@login_required
def dashboard(request):
    selected_date = request.GET.get('date')
    tasks = TaskSubmission.objects.all().order_by('-submitted_at')
    unique_dates = TaskSubmission.objects.dates('submitted_at', 'day', order='DESC')

    if selected_date:
        tasks = tasks.filter(submitted_at__date=selected_date)

    summary = tasks.values('engineer__et_id', 'engineer__name', 'task_type').annotate(count=Count('task_type'))
    # Exclude team leaders from the engineers list shown in the table
    engineers = Engineer.objects.filter(is_team_leader=False)
    task_details = tasks.select_related('engineer').prefetch_related('team_members')

    return render(request, 'dashboard.html', {
        'summary': summary,
        'engineers': engineers,
        'task_details': task_details,
        'unique_dates': unique_dates,
        'selected_date': selected_date,
    })

@login_required
def export_excel(request):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    tasks = TaskSubmission.objects.select_related('engineer').prefetch_related('team_members').all()
    if date_from and date_to:
        try:
            start_date = datetime.strptime(date_from, '%Y-%m-%d').replace(hour=0, minute=0, second=0)
            end_date = datetime.strptime(date_to, '%Y-%m-%d').replace(hour=23, minute=59, second=59)
            tasks = tasks.filter(submitted_at__range=[start_date, end_date])
        except ValueError:
            pass

    et_green = "008751"
    et_yellow = "FFC107"

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        primary_ids = tasks.values_list('engineer_id', flat=True)
        member_ids = tasks.values_list('team_members__id', flat=True)
        all_ids = {i for i in primary_ids if i is not None} | {i for i in member_ids if i is not None}
        engineers = Engineer.objects.filter(id__in=all_ids)
        for engineer in engineers:
            eng_tasks = tasks.filter(Q(engineer=engineer) | Q(team_members=engineer)).distinct()
            rows = []
            for t in eng_tasks:
                rows.append({
                    'date': t.date,
                    'shift': t.shift,
                    'reporter': t.reporter or engineer.name,
                    'location': t.location,
                    'equipment_type': t.equipment_type,
                    'problem_description': t.description,
                    'cause_of_problem': t.cause_of_problem,
                    'corrective_action': t.corrective_measure,
                    'start_time': timezone.make_naive(t.start_time, timezone.get_default_timezone()) if t.start_time and timezone.is_aware(t.start_time) else t.start_time,
                    'end_time': timezone.make_naive(t.end_time, timezone.get_default_timezone()) if t.end_time and timezone.is_aware(t.end_time) else t.end_time,
                    'time_taken': t.time_taken,
                    'status': t.status,
                    'remark': t.remark,
                })
            df = pd.DataFrame(rows)
            if df.empty:
                df = pd.DataFrame(columns=[
                    'date','shift','reporter','location','equipment_type','problem_description',
                    'cause_of_problem','corrective_action','start_time','end_time','time_taken','status','remark'
                ])
            sheet_name = f"{engineer.et_id}"[:31]
            df.to_excel(writer, sheet_name=sheet_name, index=False)

            ws = writer.sheets[sheet_name]
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color=et_green, end_color=et_green, fill_type="solid")
            thin = Side(border_style="thin", color="000000")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            center = Alignment(vertical="center", wrap_text=True)

            max_col = ws.max_column
            max_row = ws.max_row

            for col in range(1, max_col + 1):
                cell = ws.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center
                cell.border = border

            alt_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
            for row in range(2, max_row + 1):
                for col in range(1, max_col + 1):
                    cell = ws.cell(row=row, column=col)
                    if row % 2 == 0:
                        cell.fill = alt_fill
                    cell.border = border
                    cell.alignment = Alignment(wrap_text=True)

            for col in range(1, max_col + 1):
                max_length = 0
                for row in range(1, max_row + 1):
                    value = ws.cell(row=row, column=col).value
                    if value is not None:
                        length = len(str(value))
                        if length > max_length:
                            max_length = length
                ws.column_dimensions[get_column_letter(col)].width = min(max_length + 2, 80)

            ws.insert_rows(1)
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
            title_cell = ws.cell(row=1, column=1)
            title_cell.value = f"Ethiopian Airlines - Engineer {engineer.name} ({engineer.et_id})"
            title_cell.font = Font(bold=True, color="000000", size=14)
            title_cell.fill = PatternFill(start_color=et_yellow, end_color=et_yellow, fill_type="solid")
            title_cell.alignment = Alignment(horizontal="center", vertical="center")

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=tasks_by_engineer.xlsx'
    response.write(output.getvalue())
    return response

@login_required
def export_pdf(request):
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    tasks = TaskSubmission.objects.all()
    if date_from and date_to:
        try:
            start_date = datetime.strptime(date_from, '%Y-%m-%d').replace(hour=0, minute=0, second=0)
            end_date = datetime.strptime(date_to, '%Y-%m-%d').replace(hour=23, minute=59, second=59)
            tasks = tasks.filter(submitted_at__range=[start_date, end_date])
        except ValueError:
            pass

    summary = tasks.values(
        'engineer__et_id', 'engineer__name', 'task_type',
        'submitted_at', 'shift', 'location', 'equipment_type',
        'description', 'cause_of_problem', 'corrective_measure',
        'start_time', 'end_time'
    ).annotate(count=Count('task_type'))
    engineers = Engineer.objects.all()

    totals = [{'et_id': engineer.et_id, 'total': sum(count['count'] for count in summary if count['engineer__et_id'] == engineer.et_id)} for engineer in engineers]

    html = render(request, 'pdf_template.html', {'summary': summary, 'engineers': engineers, 'totals': totals}).content.decode('utf-8')

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=task_summary.pdf'
    HTML(string=html, base_url=request.build_absolute_uri()).write_pdf(response, stylesheets=['/static/css/pdf_styles.css'] if os.path.exists('/static/css/pdf_styles.css') else [])
    return response

@login_required
def download_inventory(request):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.formatting.rule import FormulaRule

    # Items: use current quantity as initial "in" (editable in sheet), start "out" at 0
    items = InventoryItem.objects.all().order_by('number')

    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory"

    # Columns: no., item, in, out, amount in stock
    headers = ["no.", "item", "in", "out", "amount in stock"]
    ws.append(headers)

    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="008751", end_color="008751", fill_type="solid")
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(vertical="center")

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    # Data rows with live formula for amount in stock = in - out
    for it in items:
        ws.append([it.number, it.item, it.quantity, 0, None])
        r = ws.max_row
        ws.cell(row=r, column=5).value = f"=C{r}-D{r}"

    # If no items, seed a few blank rows
    if ws.max_row == 1:
        for i in range(1, 4):
            ws.append([i, "", 0, 0, None])
            r = ws.max_row
            ws.cell(row=r, column=5).value = f"=C{r}-D{r}"

    last_row = ws.max_row

    # Data validation: in/out non-negative integers
    dv_int = DataValidation(type="whole", operator="greaterThanOrEqual", formula1="0", allow_blank=True)
    ws.add_data_validation(dv_int)
    dv_int.add(f"C2:C{last_row}")
    dv_int.add(f"D2:D{last_row}")

    # Conditional formatting: entire row red when out > half of in
    red_fill = PatternFill(start_color="D32F2F", end_color="D32F2F", fill_type="solid")
    half_rule = FormulaRule(formula=["$D2>($C2/2)"], stopIfTrue=False, fill=red_fill)
    ws.conditional_formatting.add(f"A2:E{last_row}", half_rule)

    # Borders, wider columns, filter, freeze
    ws.auto_filter.ref = f"A1:E{last_row}"
    ws.freeze_panes = "A2"

    for col in range(1, ws.max_column + 1):
        max_len = 0
        for row in range(1, ws.max_row + 1):
            v = ws.cell(row=row, column=col).value
            l = len(str(v)) if v is not None else 0
            if l > max_len:
                max_len = l
            ws.cell(row=row, column=col).border = border
        # Make columns a bit wider than before
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 4, 80)

    # Download
    output = io.BytesIO()
    wb.save(output)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=inventory.xlsx'
    response.write(output.getvalue())
    return response